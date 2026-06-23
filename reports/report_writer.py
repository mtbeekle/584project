from datetime import datetime
from pathlib import Path
import platform

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

from .formatting import format_report_sheet, format_summary_sheet


STANDARD_ISSUE_COLUMNS = [
    "SourceSheet",
    "RuleID",
    "Category",
    "Severity",
    "ElementType",
    "ElementID",
    "Issue",
    "Description",
    "RecommendedAction",
]


def _build_count_rows(values: pd.Series) -> list[dict[str, object]]:
    counts = (
        values.dropna()
        .astype(str)
        .value_counts()
        .sort_values(ascending=False)
    )

    return [
        {"Check": value, "Count": count}
        for value, count in counts.items()
    ]


def build_issues_log(issue_tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    issue_frames = []

    for sheet_name, dataframe in issue_tables.items():
        if dataframe.empty:
            continue

        issue_frame = dataframe.copy()
        issue_frame.insert(0, "SourceSheet", sheet_name)
        issue_frames.append(issue_frame)

    if not issue_frames:
        return pd.DataFrame(columns=STANDARD_ISSUE_COLUMNS)

    issues = pd.concat(issue_frames, ignore_index=True, sort=False)
    ordered_columns = [
        column for column in STANDARD_ISSUE_COLUMNS if column in issues.columns
    ]
    remaining_columns = [
        column for column in issues.columns if column not in ordered_columns
    ]

    return issues[ordered_columns + remaining_columns]


def build_summary_tables(
    issue_tables: dict[str, pd.DataFrame],
    issues: pd.DataFrame,
    mdb_file: Path,
    tool_version: str | None,
) -> list[tuple[str, pd.DataFrame]]:
    metadata_table = pd.DataFrame(
        [
            {
                "Label": "Report generated",
                "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            },
            {"Label": "MDB file name", "Value": mdb_file.name},
            {"Label": "Total issue count", "Value": len(issues)},
            {"Label": "Tool version", "Value": tool_version or "Unavailable"},
            {"Label": "Python version", "Value": platform.python_version()},
        ]
    )

    validation_counts = pd.DataFrame(
        [
            {"Check": sheet_name, "Count": len(dataframe)}
            for sheet_name, dataframe in issue_tables.items()
        ]
    )
    severity_counts = pd.DataFrame(
        _build_count_rows(issues["Severity"]) if "Severity" in issues.columns else [],
        columns=["Check", "Count"],
    )
    rule_counts = pd.DataFrame(
        _build_count_rows(issues["RuleID"]) if "RuleID" in issues.columns else [],
        columns=["Check", "Count"],
    )
    category_counts = pd.DataFrame(
        _build_count_rows(issues["Category"]) if "Category" in issues.columns else [],
        columns=["Check", "Count"],
    )

    return [
        ("Report Metadata", metadata_table),
        ("Validation Check Counts", validation_counts),
        ("Severity Counts", severity_counts),
        ("RuleID Counts", rule_counts),
        ("Category Counts", category_counts),
    ]


def write_summary_sheet(
    writer: pd.ExcelWriter,
    summary_tables: list[tuple[str, pd.DataFrame]],
) -> None:
    worksheet = writer.book.create_sheet(title="Summary")
    writer.sheets["Summary"] = worksheet

    table_starts = {
        "Validation Check Counts": 1,
        "Severity Counts": 4,
        "RuleID Counts": 7,
        "Category Counts": 10,
        "Report Metadata": 13,
    }

    for title, dataframe in summary_tables:
        start_column = table_starts[title]
        current_row = 1
        worksheet.cell(row=current_row, column=start_column, value=title)
        current_row += 1

        for row in dataframe_to_rows(dataframe, index=False, header=True):
            for column_index, value in enumerate(row, start=1):
                worksheet.cell(
                    row=current_row,
                    column=start_column + column_index - 1,
                    value=value,
                )
            current_row += 1

        first_header = dataframe.columns[0] if len(dataframe.columns) > 0 else None
        second_header = dataframe.columns[1] if len(dataframe.columns) > 1 else None

        if first_header == "Label":
            worksheet.column_dimensions["M"].width = 22
            worksheet.column_dimensions["N"].width = 30
        else:
            column_letter_1 = worksheet.cell(row=2, column=start_column).column_letter
            column_letter_2 = worksheet.cell(row=2, column=start_column + 1).column_letter
            worksheet.column_dimensions[column_letter_1].width = 15
            worksheet.column_dimensions[column_letter_2].width = 10

    format_summary_sheet(worksheet)


def build_issue_tables(
    missing_results: dict[str, pd.DataFrame],
    capacitor_results: dict[str, pd.DataFrame],
    fuse_results: dict[str, pd.DataFrame],
    height_results: dict[str, pd.DataFrame],
    load_results: dict[str, pd.DataFrame],
    customer_count_results: dict[str, pd.DataFrame],
    conductor_mismatch_results: dict[str, pd.DataFrame],
    incorrect_phase_results: dict[str, pd.DataFrame],
    topology_results: dict[str, pd.DataFrame],
) -> dict[str, pd.DataFrame]:
    return {
        "MissingConnectivity": missing_results["missing_connectivity"],
        "MissingLength": missing_results["missing_length"],
        "MissingPhase": missing_results["missing_phase"],
        "MissingConductor": missing_results["missing_conductor"],
        "DuplicateSections": missing_results["duplicate_sections"],
        "CapacitorIssues": capacitor_results["capacitor_issues"],
        "OpenFuses": fuse_results["open_fuses"],
        "LoopSections": topology_results["loop_sections"],
        "DisconnectedTopology": topology_results["unfed_sections"],
        "ConductorHeight": height_results["conductor_height_issues"],
        "NoConnectedKVA": load_results["no_connected_kva"],
        "CustomerCount": customer_count_results["customer_count_issues"],
        "ConductorMismatch": conductor_mismatch_results["conductor_issues"],
        "IncorrectPhases": incorrect_phase_results["incorrect_phases"],
    }


def build_report_tables(
    missing_results: dict[str, pd.DataFrame],
    capacitor_results: dict[str, pd.DataFrame],
    fuse_results: dict[str, pd.DataFrame],
    height_results: dict[str, pd.DataFrame],
    load_results: dict[str, pd.DataFrame],
    customer_count_results: dict[str, pd.DataFrame],
    conductor_mismatch_results: dict[str, pd.DataFrame],
    incorrect_phase_results: dict[str, pd.DataFrame],
    topology_results: dict[str, pd.DataFrame],
) -> dict[str, pd.DataFrame]:
    return build_issue_tables(
        missing_results,
        capacitor_results,
        fuse_results,
        height_results,
        load_results,
        customer_count_results,
        conductor_mismatch_results,
        incorrect_phase_results,
        topology_results,
    )


def build_diagnostic_tables(
    capacitor_results: dict[str, pd.DataFrame],
    topology_results: dict[str, pd.DataFrame],
) -> dict[str, pd.DataFrame]:
    return {
        "CapVoltageContext": capacitor_results.get(
            "capacitor_voltage_context",
            pd.DataFrame(),
        ),
        "TransformerLocations": capacitor_results.get(
            "transformer_locations",
            pd.DataFrame(),
        ),
        "TopologyComponents": topology_results.get(
            "topology_components",
            pd.DataFrame(),
        ),
    }


def write_validation_report(
    output_file: Path,
    mdb_file: Path,
    missing_results: dict[str, pd.DataFrame],
    capacitor_results: dict[str, pd.DataFrame],
    fuse_results: dict[str, pd.DataFrame],
    height_results: dict[str, pd.DataFrame],
    load_results: dict[str, pd.DataFrame],
    customer_count_results: dict[str, pd.DataFrame],
    conductor_mismatch_results: dict[str, pd.DataFrame],
    incorrect_phase_results: dict[str, pd.DataFrame],
    topology_results: dict[str, pd.DataFrame],
    tool_version: str | None = None,
) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    issue_tables = build_issue_tables(
        missing_results,
        capacitor_results,
        fuse_results,
        height_results,
        load_results,
        customer_count_results,
        conductor_mismatch_results,
        incorrect_phase_results,
        topology_results,
    )
    diagnostic_tables = build_diagnostic_tables(capacitor_results, topology_results)
    issues = build_issues_log(issue_tables)
    summary_tables = build_summary_tables(
        issue_tables,
        issues,
        mdb_file,
        tool_version,
    )

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        write_summary_sheet(writer, summary_tables)
        issues.to_excel(writer, sheet_name="Issues", index=False)
        format_report_sheet(writer.sheets["Issues"])

        for sheet_name, dataframe in issue_tables.items():
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
            format_report_sheet(writer.sheets[sheet_name])

        for sheet_name, dataframe in diagnostic_tables.items():
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
            format_report_sheet(writer.sheets[sheet_name])
