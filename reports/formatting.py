from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
ERROR_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
WARNING_FILL = PatternFill(fill_type="solid", fgColor="FFEB9C")
REVIEW_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
INFO_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
TOP_ALIGNMENT = Alignment(vertical="top")

COLUMN_WIDTHS = {
    "RuleID": 10,
    "Category": 18,
    "Severity": 12,
    "ElementType": 15,
    "ElementID": 18,
    "Issue": 35,
    "Description": 55,
    "RecommendedAction": 55,
}

WRAPPED_COLUMNS = {"Issue", "Description", "RecommendedAction"}


def _set_column_widths(worksheet: Worksheet) -> None:
    for cell in worksheet[1]:
        if cell.value in COLUMN_WIDTHS:
            worksheet.column_dimensions[cell.column_letter].width = COLUMN_WIDTHS[cell.value]


def _format_wrapped_columns(worksheet: Worksheet) -> None:
    header_map = {cell.value: cell.column for cell in worksheet[1]}

    for column_name in WRAPPED_COLUMNS:
        column_index = header_map.get(column_name)
        if column_index is None:
            continue

        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=column_index).alignment = WRAP_ALIGNMENT


def _top_align_data_rows(worksheet: Worksheet) -> None:
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if cell.alignment == WRAP_ALIGNMENT:
                continue
            cell.alignment = TOP_ALIGNMENT


def _format_severity_cells(worksheet: Worksheet) -> None:
    header_cells = list(worksheet[1])
    severity_column = None

    for cell in header_cells:
        if cell.value == "Severity":
            severity_column = cell.column
            break

    if severity_column is None:
        return

    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=severity_column)
        if cell.value == "Error":
            cell.fill = ERROR_FILL
        elif cell.value == "Warning":
            cell.fill = WARNING_FILL
        elif cell.value == "Review":
            cell.fill = REVIEW_FILL
        elif cell.value == "Info":
            cell.fill = INFO_FILL


def format_report_sheet(worksheet: Worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    _set_column_widths(worksheet)
    _format_wrapped_columns(worksheet)
    _top_align_data_rows(worksheet)
    _format_severity_cells(worksheet)
